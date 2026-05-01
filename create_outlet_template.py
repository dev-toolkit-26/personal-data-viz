# -*- coding: utf-8 -*-
"""파울라너 취합양식 (가로형) 생성 스크립트 - 기존 데이터 포함

실행: python create_paulaner_template.py
결과: Paulaner/파울라너_취합양식_2026.xlsx
"""
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

CURRENT_MONTH = 4  # 현재 월 (매월 업데이트)
OUTPUT_PATH = "outlet_template_2026.xlsx"

SYSTEM_OPTIONS = ['하이트시스템', '자체시스템', '하이네켄 노멀', '하이네켄 콜륨', '타사 시스템']

# ── 기존 데이터 (필요 시 kpi.html "현황 다운로드"로 Supabase에서 받아 채우기) ─
EXISTING_DATA = [
    # (region, ws, outlet, system, sub_system, m3, m4)
]

# ── 색상 ────────────────────────────────────────────────────────────
HEADER_FILL  = PatternFill("solid", fgColor="1F4E79")
INFO_FILL    = PatternFill("solid", fgColor="DDEEFF")
PAST_FILL    = PatternFill("solid", fgColor="D9E1F2")
CURRENT_FILL = PatternFill("solid", fgColor="FFE699")
FUTURE_FILL  = PatternFill("solid", fgColor="F2F2F2")
TOTAL_FILL   = PatternFill("solid", fgColor="E2EFDA")

HEADER_FONT  = Font(bold=True, color="FFFFFF", size=10)
BOLD_FONT    = Font(bold=True, size=10)
NORMAL_FONT  = Font(size=10)

thin = Side(style='thin', color="BBBBBB")
thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)

MONTHS = list(range(3, 13))
MONTH_LABELS = [f"{m}월" for m in MONTHS]
COLUMNS = ["Group", "WS", "거래처명", "System"] + MONTH_LABELS + ["합계"]
COL_WIDTHS = [12, 22, 26, 18] + [8]*10 + [10]
TOTAL_ROW = 202  # 데이터 마지막 행
BASE_COLS = 4  # Group/WS/거래처명/System (월 열 앞)

def month_fill(m):
    if m < CURRENT_MONTH:   return PAST_FILL
    if m == CURRENT_MONTH:  return CURRENT_FILL
    return FUTURE_FILL

def make_template():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "파울라너_취합"

    # ── 제목 행 ────────────────────────────────────────
    last_col_letter = get_column_letter(len(COLUMNS))
    ws.merge_cells(f"A1:{last_col_letter}1")
    c = ws["A1"]
    c.value = f"파울라너 20L 취합양식 2026  (기준월: {CURRENT_MONTH}월)"
    c.font = Font(bold=True, size=12, color="1F4E79")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24

    # ── 헤더 행 ────────────────────────────────────────
    for ci, label in enumerate(COLUMNS, 1):
        c = ws.cell(row=2, column=ci, value=label)
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = thin_border
        mi = ci - (BASE_COLS + 1)  # month index (valid for ci > BASE_COLS)
        if ci <= BASE_COLS:
            c.fill = HEADER_FILL
        elif ci == len(COLUMNS):
            c.fill = PatternFill("solid", fgColor="375623")
        elif MONTHS[mi] < CURRENT_MONTH:
            c.fill = PatternFill("solid", fgColor="2E5EA8")
        elif MONTHS[mi] == CURRENT_MONTH:
            c.fill = PatternFill("solid", fgColor="C09000")
        else:
            c.fill = PatternFill("solid", fgColor="595959")
    ws.row_dimensions[2].height = 20

    # ── 열 너비 ────────────────────────────────────────
    for ci, w in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    # ── System 드롭다운 검증 (D3:D202) ─────────────────
    sys_list = ",".join(SYSTEM_OPTIONS)
    dv = DataValidation(
        type="list",
        formula1=f'"{sys_list}"',
        allow_blank=True,
        showDropDown=False,
        showErrorMessage=True,
        errorTitle="입력 오류",
        error=f"목록에서 선택하세요: {sys_list}",
    )
    dv.sqref = f"D3:D{TOTAL_ROW}"
    ws.add_data_validation(dv)

    # ── 데이터 행 기본 스타일 적용 ─────────────────────
    for row in range(3, TOTAL_ROW + 1):
        for ci in range(1, len(COLUMNS) + 1):
            c = ws.cell(row=row, column=ci)
            c.border = thin_border
            c.font = NORMAL_FONT
            c.alignment = Alignment(horizontal="center", vertical="center")
            mi = ci - (BASE_COLS + 1)  # month index
            if ci <= BASE_COLS:
                c.fill = INFO_FILL
                c.alignment = Alignment(horizontal="left", vertical="center")
            elif ci == len(COLUMNS):
                first_col = get_column_letter(BASE_COLS + 1)
                last_col  = get_column_letter(len(COLUMNS) - 1)
                c.value = f"=SUM({first_col}{row}:{last_col}{row})"
                c.fill = TOTAL_FILL
                c.font = BOLD_FONT
            else:
                c.fill = month_fill(MONTHS[mi])

    # ── 기존 데이터 입력 ───────────────────────────────
    for i, row_data in enumerate(EXISTING_DATA):
        excel_row = 3 + i
        region, ws_name, outlet, system, _sub_sys, m3, m4 = row_data

        ws.cell(row=excel_row, column=1, value=region)
        ws.cell(row=excel_row, column=2, value=ws_name)
        ws.cell(row=excel_row, column=3, value=outlet)
        ws.cell(row=excel_row, column=4, value=system)

        month_vals = {3: m3, 4: m4}
        for m, val in month_vals.items():
            if val is not None:
                col_idx = BASE_COLS + (m - 2)  # 3월=col5, 4월=col6
                ws.cell(row=excel_row, column=col_idx, value=val)

    # ── 헤더 고정 ──────────────────────────────────────
    ws.freeze_panes = "A3"

    # ── 작성법 시트 ────────────────────────────────────
    ws2 = wb.create_sheet("작성법")
    ws2["A1"] = "파울라너 취합양식 작성법"
    ws2["A1"].font = Font(bold=True, size=12)
    notes = [
        ("A3", "■ 열 설명"),
        ("A4",  "Group"),      ("B4",  "지역: Seoul / Busan / Daejeon / Daegu / Gwangju / Jeju / NKA (대형 거래처)"),
        ("A5",  "WS"),         ("B5",  "도매상명"),
        ("A6",  "거래처명"),    ("B6",  "업장 이름"),
        ("A7",  "System"),     ("B7",  "드롭다운 선택: 하이트시스템 / 자체시스템 / 하이네켄 노멀 / 하이네켄 콜륨 / 타사 시스템"),
        ("A8",  "3월~12월"),   ("B8",  "해당 월 셀인 수량 입력"),
        ("A11", "■ 색상 안내"),
        ("A12", "노란색 열"),  ("B12", f"현재 입력 월 ({CURRENT_MONTH}월)"),
        ("A13", "파란색 열"),  ("B13", "입력 완료된 지난 월"),
        ("A14", "회색 열"),    ("B14", "아직 입력 안 할 미래 월"),
        ("A16", "■ 주의사항"),
        ("A17", ""),           ("B17", "새 업장은 입점 월부터 입력 (이전 월은 빈칸 유지)"),
        ("A18", ""),           ("B18", "업장 해지 시 해당 월부터 공란"),
        ("A19", ""),           ("B19", "업로드 시 해당 지역 전체 데이터가 교체됩니다"),
        ("A20", ""),           ("B20", "NKA(대형 거래처)는 Group에 'NKA'로 입력"),
    ]
    for addr, val in notes:
        ws2[addr] = val
        if val.startswith("■"):
            ws2[addr].font = Font(bold=True)
    ws2.column_dimensions["A"].width = 14
    ws2.column_dimensions["B"].width = 65

    wb.save(OUTPUT_PATH)
    print(f"Done: {OUTPUT_PATH}")
    print(f"  Rows populated: {len(EXISTING_DATA)}")
    print(f"  System dropdown: D3:D{TOTAL_ROW}")

if __name__ == "__main__":
    make_template()
