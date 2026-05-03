"""
update_rofo.py — Parse 'Latest RoFo' sheet of the monthly RoFo Excel and emit:
  - Rofo/rofo.sql           : Supabase patch (jsonb_set) for dashboard_snapshots id=1
  - Rofo/rofo.json (optional): debug dump of parsed values

Mirrors update_ap.py conventions. RoFo is ON-Trade Total only — no regional split.

RULE: rofo_monthly is populated only for FORECAST months of the RoFo edition.
For the "May edition" (this file), Jan-Apr are actuals (already in act_monthly /
mtd) and must NOT be overwritten by Excel's Jan-Apr columns. Only May-Dec
indices [4..11] get values; [0..3] stay null.

Usage:  python update_rofo.py
"""
import json, sys, os
from collections import defaultdict

# Windows 콘솔 cp949 회피: print의 em-dash/화살표를 안전하게 출력
try:
    sys.stdout.reconfigure(encoding="utf-8")
except Exception:
    pass

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl not installed. Run: pip install openpyxl")

HERE       = os.path.dirname(os.path.abspath(__file__))
EXCEL_PATH = os.path.join(HERE, "26Y Rofo.xlsx")
SQL_OUT    = os.path.join(HERE, "rofo.sql")
JSON_OUT   = os.path.join(HERE, "rofo.json")

CONV = 12.6263            # HL → standard cases (matches update_ap.py)
ROFO_START_IDX = 4        # 0-indexed May (edition starts here)

# Excel brand label → Dashboard brand key (from update_ap.py BRAND_MAP)
BRAND_MAP = {
    "Heineken":        "Heineken",
    "Heineken 0.0":    "Heineken 0.0",
    "Heineken Silver": "Heineken Silver",
    "Desperados":      "Desperados",
    "EdelWeiss":       "EdelWeiss",
    "Tiger":           "Tiger",
    "Tiger Radler":    "Tiger Radler",
    "Paulaner":        "Paulaner",
    "SULHWA":          "Snow",
    "Birra Moretti":   "BirraMoretti",
    # Krombacher / Apple Fox / Monteith's: not in dashboard, skipped
}

# Excel SKU label → Dashboard SKU key (where they differ; identity otherwise)
SKU_RENAME = {
    # Most labels match identity. Add overrides here if Excel labels diverge.
}

# Paulaner SKU labels in the Excel are #REF! (broken from sheet deletion).
# UNCERTAIN MAPPING: Excel hidden col J labels R51 as "2L Keg" but the only
# dashboard Paulaner SKU is `PA WHE 20000 Keg` (20L, hl_factor=0.2). If R51 is
# really a 2L keg, mapping it to the 20L dashboard SKU would inflate volume 10×.
# Until the operator confirms the actual SKU identities, leave fallback empty.
# Add e.g. `51: "PA WHE 20000 Keg"` here once verified.
PAULANER_ROW_FALLBACK = {}

# ── 1. Load Excel ───────────────────────────────────────────────
wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
ws = wb["Latest RoFo"]

# Excel layout: K=11 Team, L=12 Channel, M=13 Brand, N=14 SKU, O..Z=15..26 Jan..Dec
# Data rows: 7..58
sku_rofo = {}   # (dash_brand, dash_sku) -> [12] cases (None outside May-Dec)
skipped  = []   # (row, brand_raw, sku_raw, reason)

for r in range(7, 59):
    brand_raw = ws.cell(row=r, column=13).value
    sku_raw   = ws.cell(row=r, column=14).value

    if brand_raw is None or brand_raw == 0:
        continue
    if isinstance(sku_raw, int) and sku_raw < -2000000000:
        # #REF! error code — try fallback for Paulaner
        if brand_raw == "Paulaner" and r in PAULANER_ROW_FALLBACK:
            sku_raw = PAULANER_ROW_FALLBACK[r]
        else:
            skipped.append((r, brand_raw, "<#REF!>", "no fallback mapping"))
            continue

    dash_brand = BRAND_MAP.get(brand_raw)
    if dash_brand is None:
        skipped.append((r, brand_raw, sku_raw, "brand not in BRAND_MAP"))
        continue

    dash_sku = SKU_RENAME.get(sku_raw, sku_raw)
    key = (dash_brand, dash_sku)

    monthly = [None] * 12
    for m in range(12):
        if m < ROFO_START_IDX:
            continue   # leave null for Jan-Apr (actuals already in act_monthly)
        cell_val = ws.cell(row=r, column=15 + m).value
        if isinstance(cell_val, (int, float)):
            monthly[m] = round(float(cell_val), 2)
        else:
            monthly[m] = 0.0

    if key in sku_rofo:
        # Same dashboard SKU referenced twice in Excel — sum
        existing = sku_rofo[key]
        sku_rofo[key] = [
            (None if a is None and b is None else (a or 0) + (b or 0))
            for a, b in zip(existing, monthly)
        ]
    else:
        sku_rofo[key] = monthly

# ── 2. Convert physical cases → standard cases via hl_factor × CONV ─────
# We need each SKU's hl_factor from the dashboard. Read it from index.html /*D*/.
HTML_PATH = os.path.join(os.path.dirname(HERE), "index.html")
with open(HTML_PATH, encoding="utf-8") as f:
    html = f.read()
mp = html.index("/*D*/")
s = html.index("{", mp)
depth, e = 0, s
for i, c in enumerate(html[s:], s):
    if c == "{":
        depth += 1
    elif c == "}":
        depth -= 1
        if depth == 0:
            e = i; break
DASH = json.loads(html[s:e + 1])

ott = {b["brand"]: b for b in DASH["sku_detail"]["On Trade Total"]}
hl_factor_for = {}   # (brand, sku) -> hl_factor
known_skus    = set()
for brand, b_entry in ott.items():
    for s_entry in b_entry["skus"]:
        sku = s_entry["sku"]
        hl_factor_for[(brand, sku)] = s_entry["data"].get("hl_factor", 0.0792) or 0.0792
        known_skus.add((brand, sku))

# Filter sku_rofo to only those keys present in dashboard
sku_rofo_dash = {}
unmapped_skus = []
for key, monthly in sku_rofo.items():
    if key in known_skus:
        sku_rofo_dash[key] = monthly
    else:
        unmapped_skus.append(key)

# Convert HL → standard cases (HL × CONV)
# Excel values are in HL (verified: R59 ON Total matches dashboard monthly.act_2026 in HL).
# Standard cases convention matches existing ap_monthly storage in DASHBOARD_DATA.
sku_rofo_std = {}
for (brand, sku), monthly in sku_rofo_dash.items():
    sku_rofo_std[(brand, sku)] = [
        None if v is None else round(v * CONV, 2) for v in monthly
    ]

# ── 3. Brand-level rollup at On Trade Total ─────────────────────
brand_rofo_std = defaultdict(lambda: [0.0]*12)
brand_has_value = defaultdict(lambda: [False]*12)
for (brand, sku), monthly in sku_rofo_std.items():
    for m in range(12):
        if monthly[m] is not None:
            brand_rofo_std[brand][m] += monthly[m]
            brand_has_value[brand][m] = True

# Convert sums back to None for months with no contributing SKU
for brand in brand_rofo_std:
    brand_rofo_std[brand] = [
        round(brand_rofo_std[brand][m], 2) if brand_has_value[brand][m] else None
        for m in range(12)
    ]

# ── 4. TTL rollup (HL and standard cases) ───────────────────────
ttl_rofo_std = [0.0]*12
ttl_has_val  = [False]*12
for monthly in brand_rofo_std.values():
    for m in range(12):
        if monthly[m] is not None:
            ttl_rofo_std[m] += monthly[m]
            ttl_has_val[m] = True
ttl_rofo_std = [round(ttl_rofo_std[m], 2) if ttl_has_val[m] else None for m in range(12)]
ttl_rofo_hl  = [round(v / CONV, 2) if v is not None else None for v in ttl_rofo_std]

# ── 5. Print summary + mapping diagnostics ──────────────────────
fcst_mon = DASH.get("wk_meta", {}).get("month", 5)
print(f"FCST month from baseline: {fcst_mon}")
print(f"Skipped Excel rows: {len(skipped)}")
for r, b, s_, reason in skipped:
    print(f"  R{r}  {b!r}/{s_!r}  ({reason})")
print(f"Unmapped (brand, sku) — present in Excel but not in dashboard:")
for key in unmapped_skus:
    print(f"  {key}")
print(f"\nTTL rofo (standard cases) by month:")
for m, v in enumerate(ttl_rofo_std):
    print(f"  m{m+1:02d}: {v}")
print(f"\nTTL rofo current FCST month ({fcst_mon}):")
print(f"  HL    = {ttl_rofo_hl[fcst_mon-1]}")
print(f"  cases = {ttl_rofo_std[fcst_mon-1]}")

# ── 6. Generate rofo.sql ────────────────────────────────────────
def jsonb_set_path(*segments):
    """Build the path argument for jsonb_set."""
    return "{" + ",".join(str(s).replace("\\", "\\\\").replace("\"", "\\\"") for s in segments) + "}"

def jval(arr_or_num):
    """Format a Python value as a JSON literal for SQL."""
    return json.dumps(arr_or_num, ensure_ascii=False).replace("'", "''")

sql_lines = []
sql_lines.append("-- Generated by update_rofo.py — rolling forecast patch (May edition)")
sql_lines.append("-- Apply via Supabase SQL Editor.")
sql_lines.append("BEGIN;")
sql_lines.append("")

# Build sku_detail['On Trade Total'] index map: brand → array index
ott_brand_index = {b["brand"]: i for i, b in enumerate(DASH["sku_detail"]["On Trade Total"])}

sets = []  # list of (path_array, value_json)

# Brand-level rofo_monthly
for brand, monthly in brand_rofo_std.items():
    bi = ott_brand_index.get(brand)
    if bi is None:
        continue
    path = ["sku_detail", "On Trade Total", str(bi), "data", "rofo_monthly"]
    sets.append((path, monthly))

# SKU-level rofo_monthly
for brand, b_entry in ott.items():
    bi = ott_brand_index[brand]
    for si, s_entry in enumerate(b_entry["skus"]):
        sku = s_entry["sku"]
        monthly = sku_rofo_std.get((brand, sku))
        if monthly is None:
            # SKU exists in dashboard but not in Excel — write all-null array
            monthly = [None]*12
        path = ["sku_detail", "On Trade Total", str(bi), "skus", str(si), "data", "rofo_monthly"]
        sets.append((path, monthly))

# wk10_summary.hl.TTL.rofo (single FCST-month value, HL)
fcst_idx = fcst_mon - 1
sets.append((["wk10_summary", "hl", "TTL", "rofo_monthly"], ttl_rofo_hl))
sets.append((["wk10_summary", "cases", "BEER TTL", "rofo_monthly"], ttl_rofo_std))

# Build a single UPDATE with chained jsonb_set
# (one statement so all-or-nothing; matches update_ap.py / sku_*.sql style)
expr = "data"
for path, value in sets:
    expr = (
        "jsonb_set("
        + expr
        + ", "
        + "'" + jsonb_set_path(*path) + "'"
        + ", "
        + "'" + jval(value) + "'::jsonb"
        + ", true)"
    )

sql_lines.append("UPDATE dashboard_snapshots")
sql_lines.append(f"SET data = {expr}")
sql_lines.append("WHERE id = 1;")
sql_lines.append("")
sql_lines.append("COMMIT;")
sql_lines.append("")
sql_lines.append("-- Verification:")
sql_lines.append("-- SELECT data->'sku_detail'->'On Trade Total'->0->'data'->'rofo_monthly' FROM dashboard_snapshots WHERE id=1;")
sql_lines.append("-- SELECT data->'wk10_summary'->'hl'->'TTL'->'rofo_monthly' FROM dashboard_snapshots WHERE id=1;")

with open(SQL_OUT, "w", encoding="utf-8") as f:
    f.write("\n".join(sql_lines))
print(f"\nWrote {SQL_OUT}  ({len(sets)} jsonb_set patches)")

# ── 7. Optional debug JSON ──────────────────────────────────────
debug = {
    "fcst_month": fcst_mon,
    "ttl_rofo_hl": ttl_rofo_hl,
    "ttl_rofo_cases": ttl_rofo_std,
    "brand_rofo_cases": {b: m for b, m in brand_rofo_std.items()},
    "sku_rofo_cases": {f"{b}|{s}": m for (b, s), m in sku_rofo_std.items()},
    "skipped_rows": [{"row": r, "brand": b, "sku": str(s_), "reason": reason} for r, b, s_, reason in skipped],
    "unmapped_skus": [list(k) for k in unmapped_skus],
}
with open(JSON_OUT, "w", encoding="utf-8") as f:
    json.dump(debug, f, ensure_ascii=False, indent=2)
print(f"Wrote {JSON_OUT}")
